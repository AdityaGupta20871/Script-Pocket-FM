"""
Pocket FM Assignment — AI-Powered Book Data Extraction Pipeline
Approach: Playwright → BeautifulSoup (Part 1) → Groq LLM (Part 2 details) → Excel
Supports full pagination: Page 1 (rank 1-50) + Page 2 (rank 51-100) = 100 books
Author: Adi
"""

import asyncio
import json
import re
import time
import os
import openpyxl
from datetime import datetime
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from groq import Groq

# ─── CONFIG ──────────────────────────────────────────────────────
GROQ_API_KEY = "APNA_GROQ_KEY_YAHAN"

# Amazon Paranormal Romance Kindle Bestsellers
# Page 1 = rank 1-50 (sccl_1..50), Page 2 = rank 51-100 (sccl_1..50 again → offset +50)
BASE_URL = "https://www.amazon.com/Best-Sellers-Kindle-Store-Paranormal-Romance/zgbs/digital-text/6190484011"
BESTSELLER_PAGES = [
    (BASE_URL,           0),   # Page 1: rank offset = 0  → sccl_1 = rank 1
    (BASE_URL + "?pg=2", 50),  # Page 2: rank offset = 50 → sccl_1 = rank 51
]

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
# ─────────────────────────────────────────────────────────────────

client_ai = Groq(api_key=GROQ_API_KEY)


# ─────────────────────────────────────────────
# Fetch HTML with Playwright (US-spoofed)
# ─────────────────────────────────────────────

async def fetch_html(url: str, wait: int = 4) -> str:
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=UA,
            locale="en-US",
            timezone_id="America/New_York",
            geolocation={"latitude": 40.7128, "longitude": -74.0060},
            permissions=["geolocation"]
        )
        await ctx.add_cookies([{
            "name": "i18n-prefs",
            "value": "USD",
            "domain": ".amazon.com",
            "path": "/"
        }])
        page = await ctx.new_page()
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await page.wait_for_timeout(wait * 1000)
        html = await page.content()
        await browser.close()
        return html


# ─────────────────────────────────────────────
# PART 1: Parse one bestseller page
# rank_offset: 0 for page1, 50 for page2
# ─────────────────────────────────────────────

def parse_bestseller_page(html: str, rank_offset: int = 0) -> list:
    soup = BeautifulSoup(html, "html.parser")
    books = []

    items = soup.select("div.zg-item-immersion, li.zg-item-immersion, div[class*='zg-item']")
    if not items:
        items = soup.select("ol.zg-ordered-list li, div.zg-grid-general-faceout")

    for item in items:
        try:
            # ── RANK ─────────────────────────────────────────────
            # sccl_N exists on both pages starting from 1
            # Page 2 offset = +50 so sccl_1 on page 2 → rank 51
            rank = None
            for a in item.select("a[href]"):
                m = re.search(r"sccl_(\d+)", a.get("href", ""))
                if m:
                    rank = int(m.group(1)) + rank_offset
                    break

            # ── TITLE ────────────────────────────────────────────
            title_el = item.select_one(
                "div.p13n-sc-truncate, span.p13n-sc-truncate, "
                "a.a-link-normal span, div[class*='truncate'], span[class*='truncate']"
            )
            if not title_el:
                title_el = item.select_one("a")
            title = title_el.get_text(strip=True) if title_el else None

            # ── AUTHOR ───────────────────────────────────────────
            author_el = item.select_one(
                "a.a-size-small, span.a-size-small, "
                "div.a-row.a-size-small a, .a-color-secondary"
            )
            author = author_el.get_text(strip=True) if author_el else None

            # ── RATING ───────────────────────────────────────────
            rating = None
            rating_el = item.select_one("span.a-icon-alt")
            if rating_el:
                m = re.search(r"([\d.]+)\s+out of", rating_el.get_text())
                if m:
                    rating = float(m.group(1))

            # ── REVIEWS ──────────────────────────────────────────
            # Fix: must be a standalone number (not part of rating like "4.5")
            # Reviews are always whole numbers > 0, usually large (100s-100000s)
            # They appear as aria-label="X ratings" or inside <a> next to stars
            num_reviews = None

            # Try aria-label first (most reliable)
            for el in item.select("[aria-label]"):
                label = el.get("aria-label", "")
                m = re.search(r"([\d,]+)\s+rating", label, re.IGNORECASE)
                if m:
                    num_reviews = int(m.group(1).replace(",", ""))
                    break

            # Fallback: look for <a> that contains only a number (review count link)
            if num_reviews is None:
                for a in item.select("a[href*='customerReviews'], a[href*='#customerReviews']"):
                    rev_text = re.sub(r"[^\d]", "", a.get_text())
                    if rev_text:
                        num_reviews = int(rev_text)
                        break

            # Last fallback: span with large number (>10) that isn't a price
            if num_reviews is None:
                for span in item.select("span.a-size-small"):
                    text = span.get_text(strip=True)
                    # Skip if it looks like a rating (e.g. "4.5")
                    if re.match(r"^\d\.\d$", text):
                        continue
                    digits = re.sub(r"[^\d]", "", text)
                    if digits and int(digits) > 10:
                        num_reviews = int(digits)
                        break

            # ── PRICE ────────────────────────────────────────────
            price_el = item.select_one(
                "span.p13n-sc-price, span.a-color-price, "
                "span[class*='price'], .a-price .a-offscreen"
            )
            price = price_el.get_text(strip=True) if price_el else None

            # ── URL ──────────────────────────────────────────────
            url = None
            link_el = item.select_one("a[href*='/dp/']")
            if not link_el:
                link_el = item.select_one("a.a-link-normal[href]")
            if link_el:
                href = link_el.get("href", "")
                if href.startswith("/"):
                    href = "https://www.amazon.com" + href
                url = href.split("/ref=")[0]

            if title:
                books.append({
                    "rank": rank,
                    "title": title,
                    "author": author,
                    "rating": rating,
                    "num_reviews": num_reviews,
                    "price": price,
                    "url": url
                })
        except Exception:
            continue

    return books


# ─────────────────────────────────────────────
# PART 2: LLM extracts details from individual pages
# ─────────────────────────────────────────────

def call_llm(prompt: str) -> str:
    resp = client_ai.chat.completions.create(
        model="llama-3.1-8b-instant",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=500,
        temperature=0
    )
    return resp.choices[0].message.content.strip()


def parse_book_detail(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    relevant = ""

    for sel in ["#bookDescription_feature_div", "#productDescription", "#editorialReviews_feature_div"]:
        el = soup.select_one(sel)
        if el:
            relevant += "DESCRIPTION: " + el.get_text(strip=True)[:800] + "\n"
            break

    for sel in ["#detailBullets_feature_div", "#productDetailsTable", "#detail-bullets"]:
        el = soup.select_one(sel)
        if el:
            relevant += "DETAILS: " + el.get_text(strip=True)[:600] + "\n"
            break

    if not relevant:
        return {"description": None, "publisher": None, "publication_date": None}

    prompt = f"""Extract from this Amazon book page text:
- description (max 400 chars, the book plot/summary only)
- publisher (publisher name only, e.g. "Del Rey" or "Self-published")
- publication_date (YYYY-MM-DD format only)

Return ONLY a JSON object with exactly these 3 keys. No markdown, no explanation.
Missing fields use null.

TEXT:
{relevant[:3000]}
"""
    try:
        raw = call_llm(prompt)
        raw = re.sub(r"```json|```", "", raw).strip()
        if "{" in raw:
            raw = raw[raw.find("{"):raw.rfind("}")+1]
        return json.loads(raw)
    except Exception:
        return {"description": None, "publisher": None, "publication_date": None}


# ─────────────────────────────────────────────
# Clean data
# ─────────────────────────────────────────────

def clean_book(book: dict) -> dict:
    try:
        book["rating"] = float(book["rating"]) if book["rating"] else None
    except Exception:
        book["rating"] = None
    try:
        if isinstance(book.get("num_reviews"), str):
            cleaned = re.sub(r"[^\d]", "", book["num_reviews"])
            book["num_reviews"] = int(cleaned) if cleaned else None
        elif book.get("num_reviews"):
            book["num_reviews"] = int(book["num_reviews"])
    except Exception:
        book["num_reviews"] = None
    desc = book.get("description", "") or ""
    if "Sorry! We couldn" in desc or "amazon's home page" in desc.lower():
        book["description"] = None
    return book


# ─────────────────────────────────────────────
# Export to Excel
# ─────────────────────────────────────────────

def export_to_excel(books: list, filename: str):
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kindle Paranormal Romance"

    headers = ["Rank", "Title", "Author", "Rating", "Num Reviews",
               "Price", "URL", "Description", "Publisher", "Publication Date"]

    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for row, book in enumerate(books, 2):
        ws.cell(row=row, column=1, value=book.get("rank"))
        ws.cell(row=row, column=2, value=book.get("title"))
        ws.cell(row=row, column=3, value=book.get("author"))
        ws.cell(row=row, column=4, value=book.get("rating"))
        ws.cell(row=row, column=5, value=book.get("num_reviews"))
        ws.cell(row=row, column=6, value=book.get("price"))
        ws.cell(row=row, column=7, value=book.get("url"))
        ws.cell(row=row, column=8, value=book.get("description"))
        ws.cell(row=row, column=9, value=book.get("publisher"))
        ws.cell(row=row, column=10, value=book.get("publication_date"))

    widths = [8, 45, 25, 10, 15, 12, 55, 65, 25, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(filename)
    print(f"✅ Saved: {filename}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

async def main():
    print("🚀 Starting pipeline...\n")

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(out, exist_ok=True)

    # ── STEP 1: Fetch all pages ──────────────────────────────────
    print(f"📄 Step 1: Fetching {len(BESTSELLER_PAGES)} bestseller page(s)...")
    all_books = []
    seen_titles = set()

    for page_num, (page_url, rank_offset) in enumerate(BESTSELLER_PAGES, 1):
        print(f"\n   📃 Page {page_num} (rank offset +{rank_offset}): {page_url}")
        html = await fetch_html(page_url, wait=5)
        print(f"   HTML: {len(html):,} chars")

        with open(os.path.join(out, f"debug_page_{page_num}.html"), "w", encoding="utf-8") as f:
            f.write(html)

        page_books = parse_bestseller_page(html, rank_offset=rank_offset)
        print(f"   Raw items: {len(page_books)}")

        added = 0
        for b in page_books:
            t = (b.get("title") or "").strip()
            if t and t not in seen_titles:
                seen_titles.add(t)
                all_books.append(b)
                added += 1
        print(f"   Unique new books: {added}")

        if page_num < len(BESTSELLER_PAGES):
            print("   ⏳ Waiting 3s...")
            await asyncio.sleep(3)

    all_books.sort(key=lambda x: x.get("rank") or 999)
    print(f"\n   ✅ Total unique books: {len(all_books)}")

    if len(all_books) == 0:
        print("⚠️  No books found. Check output/debug_page_1.html")
        return

    print("\n   Sanity check (first 5 books):")
    for b in all_books[:5]:
        print(f"   Rank {b['rank']:>3}: {b['title'][:45]:<45} | Reviews: {b['num_reviews']}")

    all_books = [clean_book(b) for b in all_books]
    with open(os.path.join(out, "checkpoint_part1.json"), "w", encoding="utf-8") as f:
        json.dump(all_books, f, indent=2, ensure_ascii=False)
    print(f"\n   📌 Part 1 checkpoint saved\n")

    # ── STEP 2: Individual book pages ────────────────────────────
    print(f"📚 Step 2: Fetching details for {len(all_books)} books...")
    for i, book in enumerate(all_books):
        url = book.get("url")
        if not url:
            book.update({"description": None, "publisher": None, "publication_date": None})
            continue
        try:
            print(f"   [{i+1}/{len(all_books)}] {str(book.get('title',''))[:50]}...")
            page_html = await fetch_html(url, wait=2)
            detail = parse_book_detail(page_html)
            book.update(detail)
            book = clean_book(book)
            time.sleep(1)
        except Exception as e:
            print(f"   ⚠️  Error: {e}")
            book.update({"description": None, "publisher": None, "publication_date": None})

    with open(os.path.join(out, "checkpoint_final.json"), "w", encoding="utf-8") as f:
        json.dump(all_books, f, indent=2, ensure_ascii=False)

    # ── STEP 3: Export ───────────────────────────────────────────
    print("\n📊 Step 3: Exporting to Excel...")
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    outfile = os.path.join(out, f"kindle_paranormal_romance_{ts}.xlsx")
    export_to_excel(all_books, outfile)

    print(f"\n🎉 Done! {len(all_books)} books extracted.")
    print(f"📁 Output: {out}")


if __name__ == "__main__":
    asyncio.run(main())